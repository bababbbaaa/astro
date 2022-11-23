import openpyxl
import horoscopedb as horoscopedb
from .horoscopeproc import GetTbLen
from datetime import datetime
from horoscoperr import HandleMess

# Параметры: incol- номер колонки с данными,
# rbeg - строка файла нач.
# rend -  строка файла кон
# fName - имя файла
# TypeTable - тип таблицы, если  1, завтрашняя (постфикс _1), иначе сегодняшняя
# DeleteOld - очищать таблицу перед записью
# Листы с данными идут подряд с первого


def LoadExcMessBod(incol, rbeg, rend, fName, TypeTable, DeleteOld=False):

    try:
        cur = False
        conn = horoscopedb.ConnectDb()
        if conn is None:
            exit()
        cur = conn.cursor()

        if TypeTable == 1:
            TbName = "MessBodies_1"
        else:
            TbName = "MessBodies"

        if DeleteOld:
            # удалить строки из таблицы сообщений
            cur.execute("DELETE  FROM "+TbName)
            curid = 0
        else:
            curid = GetTbLen(conn, TbName)

        workbook = openpyxl.load_workbook(fName)
        resList = list()

        for i in range(rbeg, rend+1):
            resList.clear()
            resList.append(curid)
            curid += 1
            for nSh in range(0, 4):
                workbook.active = nSh
                worksheet = workbook.active
                resList.append(str(worksheet.cell(i, incol).value))

            cur.execute("INSERT INTO "+TbName +
                        " (ID,Col_1,Col_2,Col_3,Col_4) VALUES (?,?,?,?,?)", resList)

        conn.commit()

    except Exception as error:
        HandleMess(
            "Ошибка загрузки данных из Экселя, возможно ошибки нумерации \n"+str(error), 4, True)
        return(None)
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

# incolDate - номер колонки с датой
# incolMess - номер колонки с текстом сообщения
# rbeg - строка файла нач.
# rend -  строка файла кон
# fName - имя файла
# TypeTable - тип таблицы, если  1, завтрашняя (постфикс _1), иначе сегодняшняя
# DeleteOld - очищать таблицу перед записью


def LoadExcMessHead(incolDate, incolMess, rbeg, rend, fName, TypeTable, DeleteOld=False):

    try:
        cur = False
        conn = horoscopedb.ConnectDb()
        if conn is None:
            exit()
        cur = conn.cursor()

        if TypeTable == 1:
            TbName = "MessHeaders_1"
        else:
            TbName = "MessHeaders"

        if DeleteOld:
            cur.execute("DELETE  FROM "+TbName)  # удалить строки из таблицы
            curid = 0
        else:
            curid = GetTbLen(conn, TbName)

        workbook = openpyxl.load_workbook(fName)
        resList = list()

        for i in range(rbeg, rend+1):
            resList.clear()
            resList.append(curid)
            curid = curid+1

            workbook.active = 0
            worksheet = workbook.active

            CurrDate = str(worksheet.cell(i, incolDate).value)

            newDate = datetime.strptime(
                CurrDate, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")

            resList.append(newDate)
            resList.append(str(worksheet.cell(i, incolMess).value))

            cur.execute("INSERT INTO "+TbName +
                        "  (ID,MessDate,Header) VALUES (?,?,?)", resList)

        conn.commit()

    except Exception as error:
        HandleMess(
            "Ошибка загрузки данных из Экселя, возможно ошибки нумерации \n"+str(error), 4, True)
        return(None)
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


# загрузка самих сообщений LoadExcMessBod

# Параметры: incol- номер колонки с данными,
# rbeg - строка файла нач.
# rend -  строка файла кон
# fName - имя файла
# TypeTable - тип таблицы, если  1, завтрашняя (постфикс _1), иначе сегодняшняя
# DeleteOld - очищать таблицу перед записью
# Листы с данными идут подряд с первого

# загрузка заголовков LoadExcMessHead

# incolDate - номер колонки с датой
# incolMess - номер колонки с текстом сообщения
# rbeg - строка файла нач.
# rend -  строка файла кон
# fName - имя файла
# TypeTable - тип таблицы, если  1, завтрашняя (постфикс _1), иначе сегодняшняя
# DeleteOld - очищать таблицу перед записью


# LoadExcMessBod(2, 4, 103, "Goroskopy_na_zavtra_iyul.xlsx", 1, True)
# LoadExcMessBod(2, 4, 103, "Goroskopy_na_segodnya_iyul.xlsx", 0, True)

LoadExcMessHead(1, 2, 1, 170, "ГОРОСКОПЫ_после_корректуры_на_завтра (2).xlsx", 1, True)
LoadExcMessHead(1, 2, 1, 170, "ГОРОСКОПЫ_после корректуры (1) (1) (2).xlsx", 0, True)
